# Deprecated due to libraries not supporting
# the editing of .xlsx (Excel 2007+) format

import openpyxl
import xlsxwriter
from xlutils.copy import copy
import csv

def updateTld(fpath, startRow, sheet):
    with open(fpath) as tldCsv:
	    reader = csv.DictReader(tldCsv)
	    rowCount = startRow
	    for row in reader:
	        sheet.write(rowCount, 1, row['Top Level Domain'])
		    sheet.write(rowCount, 2, row['count'])
		    rowCount++

def updateIsp(fpath, sheet):
    r = 1
	while (sheet.cell(r,2) != ''):
	    sheet.write(r, 8, sheet.cell(r, 2).value)  # Copy ISP column
		sheet.write(r, 9, sheet.cell(r, 3).value)  # Copy count column
		r++
	with open(path) as ispCsv:
	    reader = csv.DictReader(ispCsv)
		rowCount = 1
		for row in reader:
		    sheet.write(rowCount, 2, row['ISP'])   # Import ISP column
			sheet.write(rowCount, 3, row['count']) # Import count column
			rowCount++

def updateIspServer(fpath, ssSheet, sumSheet):
    r = 1
	while (ssSheet.cell(r,2) != ''):                       
	    ssSheet.write(r, 7, ssSheet.cell(r, 2).value)  # Copy ISP column
		ssSheet.write(r, 8, ssSheet.cell(r, 3).value)  # Copy count column
	with open(path) as ispCsv:
	    reader = csv.DictReader(ispCsv)
		rowCount = 1
		for row in reader:
		    ssSheet.write(rowCount, 2, row['ISP'])         # Import ISP column
			ssSheet.write(rowCount, 3, row['Total Count']) # Import count column
			if (rowCount < 11):             # Only top 10 for summarySheet
			    sumSheet.write(0, rowCount, row['ISP'])
			    sumSheet.write(1, rowCount, row['Defacement Count'])
			    sumSheet.write(2, rowCount, row['Phishing Count'])
			    sumSheet.write(3, rowCount, row['Malware Count'])
			    sumSheet.write(5, rowCount, row['Total Count'])
			rowCount++
			
def updateIspBotnet(fpath, sheet):
    r = 1
	while (sheet.cell(r,2) != ''):                       
	    sheet.write(r, 7, sheet.cell(r, 2).value)  # Copy ISP column
		sheet.write(r, 8, sheet.cell(r, 3).value)  # Copy count column
	with open(fpath) as ispCsv:
	    reader = csv.DictReader(ispCsv)
		rowCount = 1
		for row in reader:
		    sheet.write(rowCount, 2, row['ISP'])   # Import ISP column
			sheet.write(rowCount, 3, row['count']) # Import count column
			rowCount++
			
def updateIspAll(fpath, ssSheet, sumSheet):
    r = 1
	while (ssSheet.cell(r,2) != ''):                       
	    ssSheet.write(r, 7, ssSheet.cell(r, 2).value)  # Copy ISP column
		ssSheet.write(r, 8, ssSheet.cell(r, 3).value)  # Copy count column
	with open(path) as ispCsv:
	    reader = csv.DictReader(ispCsv)
		rowCount = 1
		for row in reader:
		    ssSheet.write(rowCount, 2, row['ISP'])         # Import ISP column
			ssSheet.write(rowCount, 3, row['Total Count']) # Import count column
			if (rowCount < 11):             # Only top 10 for summarySheet
			    sumSheet.write(1, r, row['ISP'])
		        sumSheet.write(2, r, row['Defacement Count'])
		        sumSheet.write(3, t, row['Phishing Count'])
		        sumSheet.write(4, r, row['Malware Count'])
		        sumSheet.write(6, r, row['Botnet Count'])	
		        sumSheet.write(7, r, row['Total Count'])	
			rowCount++
		
				
def main():
    csvDir = '/HKSWROutput/report/'	

    df_fpath = 'monthly_report_defacement_May2015.xlsx'
    df_original = open_workbook(df_fpath)	
	deface = copy(df_original)
    
	#TODO: Update sheets 0, 1, 4, 7 of servers
	updateTld(csvDir + 'DefacementTld.csv', 3, deface.get_sheet(2))
    updateTld(csvDir + 'PhishingTld.csv', 1, deface.get_sheet(5))
    updateTld(csvDir + 'MalwareTld.csv', 1, deface.get_sheet(8))
    
	updateIsp(csvDir + 'ISPDefacement.csv', deface.get_sheet(3))
    updateIsp(csvDir + 'ISPPhishing.csv', deface.get_sheet(6))
    updateIsp(csvDir + 'ISPMalware.csv', deface.get_sheet(9))
	
	updateIspServer(csvDir + 'ISPServerAll.csv', 
	                    deface.get_sheet(10),
	                    deface.get_sheet(11))
	
    deface.save(df_fpath)
    
	bn_fpath = 'botnet_report_monthly_May2015.xlsx'
	bn_original = open_workbook(bn_fpath)
	botnet = copy(bn_original)
	
	#TODO: Update sheets 0, 2 of botnet
    updateIspBotnet(csvDir + 'ISPBotnetts.csv', botnet.get_sheet(1))
	updateIspAll(csvDir + 'ISPAll.csv', botnet.get_sheet(3), 
	                botnet.get_sheet(4))
					
	botnet.save(bn_fpath)
	
if __name__ == "__main__":	
    main()	

